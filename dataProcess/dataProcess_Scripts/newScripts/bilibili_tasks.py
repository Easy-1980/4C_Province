from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_COLUMNS = ("bvid", "opera", "province", "status")
VALID_STATUS = {"processed", "unprocessed"}


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_bilibili_tasks_json(input_path: Path, output_path: Path) -> None:
    workbook = load_workbook(filename=input_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            raise ValueError("bilibili_tasks.xlsx 为空，缺少表头。")

        header_map: dict[str, int] = {}
        for idx, name in enumerate(header_row):
            key = _normalize_cell(name).lower()
            if key:
                header_map[key] = idx

        missing = [col for col in REQUIRED_COLUMNS if col not in header_map]
        if missing:
            raise ValueError(f"bilibili_tasks.xlsx 缺少必需列：{', '.join(missing)}")

        processed_items: list[dict[str, str]] = []
        unprocessed_items: list[dict[str, str]] = []

        for row_index, row in enumerate(rows, start=2):
            bvid = _normalize_cell(row[header_map["bvid"]] if header_map["bvid"] < len(row) else "")
            if not bvid:
                continue

            status = _normalize_cell(row[header_map["status"]] if header_map["status"] < len(row) else "").lower()
            if status not in VALID_STATUS:
                print(f"[bilibili_tasks] warning: row {row_index} has invalid status '{status}', skipped")
                continue

            item = {
                "bvid": bvid,
                "opera": _normalize_cell(row[header_map["opera"]] if header_map["opera"] < len(row) else ""),
                "province": _normalize_cell(row[header_map["province"]] if header_map["province"] < len(row) else ""),
                "status": status,
            }
            if status == "processed":
                processed_items.append(item)
            else:
                unprocessed_items.append(item)

        payload = {
            "summary": {
                "total": len(processed_items) + len(unprocessed_items),
                "processed": len(processed_items),
                "unprocessed": len(unprocessed_items),
            },
            "processed": processed_items,
            "unprocessed": unprocessed_items,
        }

        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    finally:
        workbook.close()


if __name__ == "__main__":
    project_root = Path(__file__).resolve().parents[3]
    input_file = project_root / "dataProcess" / "rawData" / "bilibili_tasks.xlsx"
    output_file = project_root / "dataProcess" / "output" / "bilibili_tasks.json"
    build_bilibili_tasks_json(input_path=input_file, output_path=output_file)
    print(f"[bilibili_tasks] generated: {output_file}")
